"""
Generate an Excel file that lays out a Vietnamese retail invoice similar to the provided sample.

Quick start:
    python generate_invoice_excel.py

This will create "hoa_don_ban_le.xlsx" in the current folder with example data.

You can also import build_invoice_excel(...) in your own code.

Notes:
- Number formatting uses "#,##0" so thousands display with separators (Excel will localize
  the symbol based on your OS/Excel locale). Adjust if you want fixed Vietnamese format.
- All text and blocks are fully editable in Excel after generation.
"""
import os
from datetime import datetime
from typing import  Dict, Optional

from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins # pylint: disable=ungrouped-imports
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string

from tools.common import TableAttribute, CustomerAttribute
from tools.utils import load_json, add_image_fit_cell, export_xlsx_to_pdf

CONFIG_PATH = os.environ['CONFIG_PATH']
SHOP_INFO_PATH = os.environ['SHOP_INFO_PATH']

THIN = Side(style="thin", color="000000")
MED  = Side(style="medium", color="000000")
THICK = Side(style="thick", color="000000")
DASH = Side(style='dashed', color="000000")
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")

DEFAULT_COLUMNS_WIDTH = {
    "A": 5,   # STT
    "B": 34,  # Tên mặt hàng
    "C": 16,  # Đơn vị tính
    "D": 14,  # Số lượng
    "E": 18,  # Đơn giá
    "F": 10,  # Thành tiền 1
    "G": 15   # Thành tiền 2
}

DEFAULT_ROW_HEIGHT = {
    '3': 45,
    '4': 30
}

Item = Dict[str, Optional[str]]  # name, unit, qty, price

class InvoiceBuilder():
    """ Invoicce Builder Class """
    def __init__(self):
        self.customer_name = None
        self.config = load_json(CONFIG_PATH)
        self.shop_info = load_json(SHOP_INFO_PATH)

        # Create workbook
        self.wb = Workbook()
        self.ws = None

    def _sheet_init(self):
        self.ws = self.wb.create_sheet("HoaDon")
        self.ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4)
        self.ws.sheet_properties.pageSetUpPr.fitToPage = True
        self.ws.page_setup.fitToWidth = 1
        self.ws.page_setup.fitToHeight = 0
        self.ws.sheet_view.showGridLines = False
        self._set_col_widths(DEFAULT_COLUMNS_WIDTH)
        self._set_row_heights(DEFAULT_ROW_HEIGHT)


    def _set_col_widths(self, widths: Dict[str, float]):
        for col, width in widths.items():
            self.ws.column_dimensions[col].width = width

    def _set_row_heights(self, heights: Dict[str, float]):
        for row, height in heights.items():
            self.ws.row_dimensions[int(row)].height = height

    def _merge(self, cell_range: str, value: str = "", font: Optional[Font] = None,
            align: Alignment = Alignment(horizontal="left", vertical="center")):
        self.ws.merge_cells(cell_range)
        cell = self.ws[cell_range.split(":")[0]]
        cell.value = value
        cell.alignment = align
        if font:
            cell.font = font
        return cell

    def _label(self, ws, cell: str, text: str, bold=False, italic=False, size=11,
            align="left"):
        c = ws[cell]
        c.value = text
        c.font = Font(bold=bold, italic=italic, size=size)
        c.alignment = Alignment(horizontal=align, vertical="center")
        return c

    def _draw_line(self, row:int, start_col:str, end_col:str, line_style:Side=THIN):
        start_idx = column_index_from_string(start_col)
        end_idx = column_index_from_string(end_col)
        for col in range(start_idx, end_idx+1):  # A1:E1
            self.ws.cell(row=row, column=col).border = Border(bottom=line_style)

    def _save_file(self, invoice_id:str):
        export_path = os.path.join(self.config.export_folder, f"{invoice_id}.xlsx")
        os.makedirs(self.config.export_folder, exist_ok=True)

        # Save xlsx file
        self.wb.save(export_path)

        # Export to pdf
        pdf_path = export_xlsx_to_pdf(export_path, remove_xlsx=True)
        return pdf_path

# -------------------------------
# Core builder
# -------------------------------

    def _build_shop_info(self):
        left_start = 'A'
        left_end = 'D'
        right_start = 'E'
        right_end = 'G'
        normal_size = 12
        middle_size = 16
        big_size = 25

        row = 2
        self._merge(
            f"{left_start}{row}:{left_end}{row}",
            value=self.shop_info.shop_title,
            font=Font(bold=True, size=normal_size),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

        row = 3
        self._merge(
            f"{left_start}{row}:{left_end}{row}",
            value=self.shop_info.shop_name,
            font=Font(bold=True, size=big_size),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

        row = 4
        self._merge(
            f"{left_start}{row}:{left_end}{row}",
            value=self.shop_info.shop_address,
            font=Font(bold=False, size=normal_size),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

        row = 5
        self._merge(
            f"{left_start}{row}:{left_end}{row}",
            value=self.shop_info.shop_phone,
            font=Font(bold=True, size=normal_size),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

        row = 6
        self._merge(
            f"{left_start}{row}:{left_end}{row}",
            value=self.shop_info.shop_bank,
            font=Font(bold=True, size=normal_size),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

        row = 7
        self._merge(
            f"{left_start}{row}:{left_end}{row}",
            value=self.shop_info.shop_owner,
            font=Font(bold=True, size=normal_size),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

        row_start = 3
        row_end = 4
        self._merge(
            f"{right_start}{row_start}:{right_end}{row_end}",
            value=self.shop_info.shop_specialty,
            font=Font(bold=True, size=middle_size),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

        row = 6
        col = 'F'
        self._merge(
            f"{right_start}{row}:{col}{row}",
            value=self.shop_info.invoice_title,
            font=Font(bold=True, size=normal_size),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

        self._merge(
            f"{right_end}{row}:{right_end}{row}",
            value=self.shop_info.invoice_number,
            font=Font(bold=True, size=normal_size),
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
        )

        row = 8
        self._draw_line(row, left_start, right_end)
        return row

    def _build_customer_info(self, start_row:int, customer_data:dict):
        left_start = 'A'
        left_end = 'B'
        right_start = 'C'
        right_end = 'G'
        normal_size = 12

        start_row = start_row + 1
        for index, (key, item) in enumerate(customer_data.items()):
            row = start_row + index
            self._merge(
                f"{left_start}{row}:{left_end}{row}",
                value=key,
                font=Font(bold=True, size=normal_size),
                align=Alignment(horizontal="center", vertical="center", wrap_text=True),
            )

            self._merge(
                f"{right_start}{row}:{right_end}{row}",
                value=item,
                font=Font(bold=True, size=normal_size),
                align=Alignment(horizontal="center", vertical="center", wrap_text=True),
            )

            self._draw_line(row, right_start, right_end, DASH)

        row += 1
        self._draw_line(row, left_start, right_end)
        return row

    def _build_invoice(self, start_row:int, invoice_data:dict|list):
        """ Buy main invoice """

        total_size = 15
        header_align = Alignment(horizontal="center", vertical="center")
        item_align = Alignment(horizontal="center", vertical="center")
        money_align = Alignment(horizontal="right", vertical="center")

        # Table header
        start_col = 1 # A
        current_row = start_row + 2
        headers = ["STT", "Tên mặt hàng", "Đơn vị tính", "Số lượng", "Đơn giá", "Thành tiền"]
        for index, header in enumerate(headers[:-1], start=start_col):
            cell = self.ws.cell(row=current_row, column=index, value=header)
            cell.font = Font(bold=True)
            cell.alignment = header_align

        s_col = get_column_letter(start_col + len(headers)-1)
        e_col = get_column_letter(start_col + len(headers))
        cell = self._merge(
            f"{s_col}{current_row}:{e_col}{current_row}",
            value=headers[-1],
            font=Font(bold=True),
            align=header_align,
        )

        # Table body
        current_row = current_row + 1
        total_formula_cells = []
        for i, item in enumerate(invoice_data, start=1):
            # STT
            self.ws.cell(
                row=current_row,
                column=start_col,
                value=i
            ).alignment = item_align

            # Tên mặt hàng
            self.ws.cell(
                row=current_row,
                column=start_col+1,
                value=item[TableAttribute.NAME.value]
            ).alignment = item_align

            # Đơn vị tính
            self.ws.cell(
                row=current_row,
                column=start_col+2,
                value=item[TableAttribute.TYPE.value]
            ).alignment = item_align

            # Số lượng
            qty_cell = self.ws.cell(
                row=current_row,
                column=start_col+3,
                value=int(item[TableAttribute.QUANTITY.value])
            )
            qty_cell.number_format = "#,##0"
            qty_cell.alignment = item_align

            # Đơn giá
            price_cell = self.ws.cell(
                row=current_row,
                column=start_col+4,
                value=int(item[TableAttribute.PRICE.value])
            )
            price_cell.number_format = "#,##0"
            price_cell.alignment = money_align

            # Thành Tiền
            s_col = get_column_letter(start_col+5)
            e_col = get_column_letter(start_col+6)
            amt_cell = self._merge(
                f"{s_col}{current_row}:{e_col}{current_row}",
                value=f"=D{current_row}*E{current_row}",
                align=money_align,
            )
            amt_cell.number_format = "#,##0"
            total_formula_cells.append(amt_cell.coordinate)

            current_row += 1

        # Tổng Tiền
        s_col = get_column_letter(start_col)
        e_col = get_column_letter(start_col+4)
        self._merge(
            f"{s_col}{current_row}:{e_col}{current_row}",
            value="Tổng Tiền:",
            font=Font(bold=True, size=total_size),
            align=money_align,
        )

        s_col = get_column_letter(start_col+5)
        e_col = get_column_letter(start_col+6)
        total_cell = self._merge(
            f"{s_col}{current_row}:{e_col}{current_row}",
            value=f"=SUM({','.join(total_formula_cells)})" if total_formula_cells else 0,
            font=Font(bold=True, size=total_size),
            align=money_align,
        )
        total_cell.number_format = "#,##0"

        # Border
        for col in range(start_col, start_col+7):
            self.ws.cell(row=start_row+3, column=col).border = Border(top=THICK)
        for col in range(start_col, start_col+7):
            self.ws.cell(row=current_row, column=col).border = Border(top=THICK)

        return current_row

    def _build_confirm(self, start_row:int):
        item_align = Alignment(horizontal="center", vertical="center")

        row = start_row + 2
        buyer_start = 'A'
        buyer_end = 'B'
        confirm_start = 'C'
        confirm_end = 'D'
        seller_start = 'E'
        seller_end = 'G'

        self._merge(
            f'{seller_start}{row}:{seller_end}{row}',
            value=datetime.now().strftime("Ngày %d, Tháng %m, Năm %Y"),
            font=Font(bold=False),
            align=item_align
        )

        row += 1
        self._merge(
            f'{buyer_start}{row}:{buyer_end}{row}',
            value="Đơn vị mua hàng",
            font=Font(bold=True),
            align=item_align
        )
        self._merge(
            f'{confirm_start}{row}:{confirm_end}{row}',
            value="ĐÃ NHẬN ĐỦ TIỀN",
            font=Font(bold=True),
            align=item_align
        )
        self._merge(
            f'{seller_start}{row}:{seller_end}{row}',
            value="Đơn vị bán hàng",
            font=Font(bold=True),
            align=item_align
        )

        row += 1
        self._merge(
            f'{buyer_start}{row}:{buyer_end}{row}',
            value="(Kí và ghi rõ họ tên)",
            font=Font(bold=False),
            align=Alignment(horizontal="center", vertical="top")
        )
        self.ws.row_dimensions[int(row)].height = 50

        self._merge(
            f'{seller_start}{row}:{seller_end}{row}',
            value="",
            font=Font(bold=False),
            align=item_align
        )
        add_image_fit_cell(
            self.ws,
            self.shop_info.sign_img,
            f'{seller_start}{row}',
            f'{seller_end}{row}',
            fit_inside=True)

        row += 1
        self._merge(
            f'{seller_start}{row}:{seller_end}{row}',
            value=self.shop_info.shop_owner,
            font=Font(bold=True),
            align=item_align
        )
        return row

    def build(self, invoice_id:str, invoice_data:list, customer_data:list):
        """ Build invoice and save under excel file """
        for sheet in self.wb.sheetnames:
            del self.wb[sheet]
        self.customer_name = customer_data[CustomerAttribute.NAME.value]
        self._sheet_init()
        row = self._build_shop_info()
        row = self._build_customer_info(start_row=row, customer_data=customer_data)
        row = self._build_invoice(start_row=row, invoice_data=invoice_data)
        row = self._build_confirm(start_row=row)
        self._save_file(invoice_id)
